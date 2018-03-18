from urllib.request import urlopen
from bs4 import BeautifulSoup
from datetime import datetime
import openpyxl
# URLの取得
html = urlopen("http://www.traders.co.jp/domestic_stocks/invest_tool/futures/futures_top.asp")
# htmlをBeautifulSoupで読み込み
bsObj = BeautifulSoup(html, "html.parser")

"""表の数字の取得"""
#テーブルを指定
table = bsObj.findAll("td", {"class":"list_cell"}) # tdかつclassがlist_cellのもののみを抽出

price = [] # 配列の準備
for i in range(len(table)): # for文は配列の長さ分
    price.append(table[i].getText().strip()) # getTextで<><>の間の文字を抽出し、strip()で改行を削除
"""表の数字の取得完了"""

"""証券会社が何列目に存在するかチェック"""
# Classの絞り込み
elems = bsObj.select('.list_broker')
# 指定した証券会社の行数をカウント
for i in range(len(elems)):
    text = elems[i].getText()
    if text.strip() == "ゴールドマン":
        goldman_row = i # 指定した会社名が何行名に存在するか記録する
    if text.strip() == "ＡＢＮアムロ":
        amuro_row = i # 指定した会社名が何行名に存在するか記録する
"""証券会社が何列目に存在するかチェックの終了"""

"""xlsxの指定位置に書き込んでいく"""
# 指定したxlsxを開く
wb = openpyxl.load_workbook('goldman_sakimono.xlsx', data_only = True)
# シート指定
ws = wb.get_sheet_by_name('Sakimono')
# 最終行の行番号の取得
last_row = ws.max_row
# 現在の時刻を年、月、日、時、分、秒で取得
time_ = datetime.now().strftime("%Y/%m/%d %H:%M:%S")
# 1列目に時間を挿入
ws.cell(row=last_row + 1, column=1, value=time_)
# 実際に数値を記入していく
if 'goldman_row' in globals(): # ファイルが存在しない場合の分岐
    for i in range(2):
        ws.cell(row=last_row + 1, column = i + 2, value=price[(goldman_row)*9+i])
    else:
        pass # ファイルがない時は何もしない
if 'goldman_row' in globals(): # ファイルが存在しない場合の分岐
    for i in range(2):
        ws.cell(row=last_row + 1, column = i + 4, value=price[(goldman_row)*9+i+3])
    else:
        pass # ファイルがない時は何もしない
if 'goldman_row' in globals(): # ファイルが存在しない場合の分岐
    for i in range(2):
        ws.cell(row=last_row + 1, column = i + 6, value=price[(goldman_row)*9+i+6])
    else:
        pass # ファイルがない時は何もしない

# ファイルの保存
wb.save(filename='goldman_sakimono.xlsx')
"""xlsxの指定位置への書き込み終了"""

"""以下同様のコードの別証券番号のためコメント略"""
"""ABNアムロ"""
wb = openpyxl.load_workbook('abnamuro_sakimono.xlsx', data_only = True)
ws = wb.get_sheet_by_name('Sakimono')
last_row = ws.max_row
time_ = datetime.now().strftime("%Y/%m/%d %H:%M:%S")
ws.cell(row=last_row + 1, column=1, value=time_)
if 'goldman_row' in globals():
    for i in range(2):
        ws.cell(row=last_row + 1, column = i + 2, value=price[(amuro_row)*9+i])
    else:
        pass
if 'goldman_row' in globals():
    for i in range(2):
        ws.cell(row=last_row + 1, column = i + 4, value=price[(amuro_row)*9+i+3])
    else:
        pass
if 'goldman_row' in globals():
    for i in range(2):
        ws.cell(row=last_row + 1, column = i + 6, value=price[(amuro_row)*9+i+6])
    else:
        pass
wb.save(filename='abnamuro_sakimono.xlsx')